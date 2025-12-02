import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "raw"
CENSUS_XLSX = RAW / "census.xlsx"
OUT_DIR = ROOT.parent.parent.parent / "generated-data"
OUT_CSV = OUT_DIR / "census.csv"

COLUMN_PROVINCE = 1
COLUMN_DISTRICT = 2
COLUMN_LOCAL = 3
COLUMN_SEX = 4
COLUMN_CASTE = 5
COLUMN_VALUE = 6
START_ROW = 35


def cell_str(ws, row_idx0, col_idx0):
    cell = ws.cell(row=row_idx0 + 1, column=col_idx0 + 1)
    v = cell.value

    if v is None:
        return None

    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None

    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    if isinstance(v, bool):
        return str(v)

    s = str(v).strip()

    return s if s != "" else None


def cell_num(ws, row_idx0, col_idx0):
    cell = ws.cell(row=row_idx0 + 1, column=col_idx0 + 1)
    v = cell.value

    if v is None:
        return None

    if isinstance(v, (int, float)):
        return float(v)

    if isinstance(v, str):
        try:
            return float(v)
        except Exception:
            return None

    return None


def wrangle(census_xlsx_path: Path):
    wb = load_workbook(filename=str(census_xlsx_path), data_only=True)
    sheet = wb.worksheets[0]

    last_row0 = sheet.max_row - 1
    raws = []

    for r in range(START_ROW, last_row0 + 1):
        province = cell_str(sheet, r, COLUMN_PROVINCE)
        district = cell_str(sheet, r, COLUMN_DISTRICT)
        locality = cell_str(sheet, r, COLUMN_LOCAL)
        sex = cell_str(sheet, r, COLUMN_SEX)
        caste = cell_str(sheet, r, COLUMN_CASTE)
        value = cell_num(sheet, r, COLUMN_VALUE)

        if all(v is None for v in [province, district, locality, sex, caste, value]):
            continue

        raws.append(
            {
                "rownum": r,
                "province": province,
                "district": district,
                "locality": locality,
                "sex": sex,
                "caste_ethnicity": caste,
                "value": value,
            }
        )

    current_province = None
    for i, raw in enumerate(raws):
        if raw["province"] is not None:
            current_province = raw["province"]

        raws[i]["province"] = current_province

    current_district = None
    last_province_for_district = None
    for i, raw in enumerate(raws):
        if raw["province"] != last_province_for_district:
            current_district = None
            last_province_for_district = raw["province"]

        if raw["district"] is not None:
            current_district = raw["district"]

        raws[i]["district"] = current_district

    current_local = None
    last_province_for_local = None
    last_district_for_local = None

    for i, raw in enumerate(raws):
        if (
            raw["province"] != last_province_for_local
            or raw["district"] != last_district_for_local
        ):
            current_local = None
            last_province_for_local = raw["province"]
            last_district_for_local = raw["district"]

        if raw["locality"] is not None:
            current_local = raw["locality"]

        raws[i]["locality"] = current_local

    with_geo = [r for r in raws if r["province"] and r["district"] and r["locality"]]
    current_sex = None
    last_local_key = None

    for raw in with_geo:
        key = f"{raw['province']}|{raw['district']}|{raw['locality']}"

        if key != last_local_key:
            current_sex = None
            last_local_key = key

        if raw["sex"] is not None:
            current_sex = raw["sex"]

        raw["sex"] = current_sex

    finalized = [
        r
        for r in with_geo
        if r["sex"] is not None
        and r["caste_ethnicity"] is not None
        and r["value"] is not None
        and not (
            isinstance(r["locality"], str)
            and r["locality"].strip().lower() == "institutional"
        )
    ]

    wb.close()
    return finalized


if __name__ == "__main__":
    if not CENSUS_XLSX.exists():
        raise SystemExit(f"census xlsx not found at {CENSUS_XLSX}")

    rows = wrangle(CENSUS_XLSX)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    rows_csv = [{k: v for k, v in r.items() if k != "rownum"} for r in rows]

    with OUT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "province",
                "district",
                "locality",
                "sex",
                "caste_ethnicity",
                "value",
            ],
        )

        writer.writeheader()
        writer.writerows(rows_csv)
